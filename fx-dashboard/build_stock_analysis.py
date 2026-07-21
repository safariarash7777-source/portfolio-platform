# -*- coding: utf-8 -*-
"""
ساخت خودکار و بازتولیدپذیرِ فایل «تحلیل کامل بورس» از فایل خامِ Screener Plus.

هدف: با هر آپدیتِ داده، فقط همین اسکریپت را اجرا کن تا فایلِ تحلیل و در نتیجه
تب «بورس» داشبورد به‌روز شود — بدون هیچ دخالت دستی.

اجرا:
    python build_stock_analysis.py [مسیر فایل خام screener]

اگر آرگومان ندهی، از RAW_DEFAULT پایین استفاده می‌شود.

خروجی (کنار همین اسکریپت):
    تحلیل_کامل_بورس_<تاریخ>.xlsx     (نسخهٔ آرشیویِ تاریخ‌دار)
    تحلیل_کامل_بورس_جاری.xlsx        (کپی ثابت؛ داشبورد همیشه از این می‌خواند)
"""
import os
import re
import sys
import shutil

import numpy as np
import pandas as pd

# ──────────────────────────────────────────────────────────────────────────────
# پیکربندی مسیرها
# ──────────────────────────────────────────────────────────────────────────────
HERE = os.path.dirname(os.path.abspath(__file__))
# مسیر پیش‌فرضِ فایل خام (در صورت ندادن آرگومان) — برای فایل‌های آینده قابل‌تغییر
RAW_DEFAULT = r"C:\Users\arsa\Downloads\screener-plus-1405-03-21.xlsx"
CURRENT_NAME = "تحلیل_کامل_بورس_جاری.xlsx"   # نامِ ثابتی که داشبورد می‌خواند

# ساختارِ دردسرسازِ فایل خام: هدر واقعی در ردیف ۸ → header=7
HEADER_ROW = 7

# ستون‌های خام مورد استفاده و نگاشتِ آن‌ها به نام فارسیِ موردنیازِ داشبورد
# (نام‌ها دقیقاً همان‌هایی است که app.py و data_sources.py انتظار دارند)
RENAME = {
    "Symbol": "نماد", "Company": "شرکت", "Sector": "سکتور", "Industry": "صنعت",
    "Price": "قیمت", "PtoE_ttm": "PE", "PtoB": "PB", "PtoS": "PS",
    "NI_1YRGR": "رشد_سود", "Rev_1YrGR": "رشد_درآمد",
    "ROE": "ROE", "ROA": "ROA", "NPMargin": "حاشیه_خالص", "OPMargin": "حاشیه_عملیاتی",
    "Ret_1Yr": "بازده_1y", "Ret_3M": "بازده_3m", "Ret_YTD": "بازده_ytd", "DivYield": "بازده_تقسیمی",
    "NetFx_IRR": "خالص_ارزی", "DebtRatio": "نسبت_بدهی",
}
# ستون‌هایی که در فایل خام «کسر/نسبت» هستند و باید ×۱۰۰ شوند تا عددِ درصد گردند
PCT_COLS = ["رشد_سود", "رشد_درآمد", "ROE", "ROA", "حاشیه_خالص", "حاشیه_عملیاتی",
            "بازده_1y", "بازده_3m", "بازده_ytd", "بازده_تقسیمی", "نسبت_بدهی"]
# شاخص‌هایی که در سطح صنعت/سکتور میانه گرفته می‌شوند
IND_MEDIAN_COLS = ["PE", "PB", "PS", "رشد_سود", "رشد_درآمد", "ROE", "ROA",
                   "حاشیه_خالص", "حاشیه_عملیاتی", "بازده_1y", "بازده_3m", "بازده_ytd", "بازده_تقسیمی"]

# ترتیب نهاییِ ستون‌ها (داشبورد به این ترتیب/نام‌ها وابسته است)
STOCK_COLS = ["نماد", "شرکت", "سکتور", "صنعت", "قیمت", "PE", "PB", "PS",
              "رشد_سود", "رشد_درآمد", "ROE", "ROA", "حاشیه_خالص", "حاشیه_عملیاتی",
              "بازده_1y", "بازده_3m", "بازده_ytd", "بازده_تقسیمی",
              "خالص_ارزی", "نسبت_بدهی", "ارزش_بازار_همت"]
INDUSTRY_COLS = ["صنعت", "رتبه", "امتیاز", "وزن_بازار٪", "تعداد", "PE", "PB", "PS",
                 "رشد_سود", "رشد_درآمد", "ROE", "ROA", "حاشیه_خالص", "حاشیه_عملیاتی",
                 "بازده_1y", "بازده_3m", "بازده_ytd", "بازده_تقسیمی", "دلاری٪"]
SECTOR_COLS = ["سکتور", "وزن٪", "PE", "ROE", "رشد_سود", "بازده_1y", "تعداد"]

MIN_STOCKS_FOR_SCORE = 4   # صنایع با کمتر از ۴ شرکت امتیاز نمی‌گیرند (بی‌معنیِ آماری)


# ──────────────────────────────────────────────────────────────────────────────
# گام ۰ — ابزارها
# ──────────────────────────────────────────────────────────────────────────────
def safe_num(series: pd.Series) -> pd.Series:
    """
    تبدیل امن به عدد: str → حذف کامای هزارگان → strip → '-'/خالی → NaN →
    to_numeric(errors='coerce') → بی‌نهایت‌ها به NaN.
    """
    s = series.astype(str).str.replace(",", "", regex=False).str.strip()
    s = s.replace({"-": np.nan, "": np.nan, "nan": np.nan, "None": np.nan, "NaN": np.nan})
    return pd.to_numeric(s, errors="coerce").replace([np.inf, -np.inf], np.nan)


def zscore(s: pd.Series) -> pd.Series:
    """z-score استاندارد با انحراف معیار نمونه‌ای (ddof=1)."""
    sd = s.std(ddof=1)
    if not sd or pd.isna(sd):
        return pd.Series(0.0, index=s.index)
    return (s - s.mean()) / sd


def parse_date_token(fname: str):
    """از نام فایلِ خام، تاریخ شمسی yyyymmdd را درمی‌آورد → ('1405/03/21', '1405-03-21')."""
    digits = "".join(re.findall(r"\d", os.path.basename(fname)))
    m = re.search(r"(13|14)\d{6}", digits)
    if not m:
        return "نامشخص", "نامشخص"
    d = m.group(0)
    return f"{d[:4]}/{d[4:6]}/{d[6:8]}", f"{d[:4]}-{d[4:6]}-{d[6:8]}"


# ──────────────────────────────────────────────────────────────────────────────
# گام ۱ + ۲ — خواندن، پاکسازی و ساخت جدول سهام
# ──────────────────────────────────────────────────────────────────────────────
def build_stocks(raw_path: str) -> pd.DataFrame:
    raw = pd.read_excel(raw_path, header=HEADER_ROW)
    raw = raw[raw["Symbol"].notna()].copy()           # حذف ردیف‌های بدون نماد

    d = pd.DataFrame()
    d["نماد"] = raw["Symbol"].astype(str).str.strip()
    d["شرکت"] = raw["Company"].astype(str).str.strip()
    d["سکتور"] = raw["Sector"].astype(str).str.strip()
    d["صنعت"] = raw["Industry"].astype(str).str.strip()
    # ستون‌های عددی با تبدیل امن
    for raw_col, fa in RENAME.items():
        if fa in ("نماد", "شرکت", "سکتور", "صنعت"):
            continue
        d[fa] = safe_num(raw[raw_col])

    # outlier در P/E: ≤۰ (زیان‌ده) یا >۱۰۰ (نویز) → NaN
    d["PE"] = d["PE"].where((d["PE"] > 0) & (d["PE"] <= 100))

    # درصدها ×۱۰۰
    for c in PCT_COLS:
        d[c] = d[c] * 100

    # ارزش بازار به همت (÷۱e۱۳)
    d["ارزش_بازار_همت"] = (safe_num(raw["MarketCap"]) / 1e13).round(2)

    # گردکردنِ ملایم برای خوانایی (روی بازتولیدِ رتبه/شمارش اثری ندارد)
    for c in ["PE", "PB", "PS"]:
        d[c] = d[c].round(2)
    for c in PCT_COLS:
        d[c] = d[c].round(1)
    d["قیمت"] = d["قیمت"].round(0)

    d = d[STOCK_COLS].sort_values("ارزش_بازار_همت", ascending=False).reset_index(drop=True)
    return d


# ──────────────────────────────────────────────────────────────────────────────
# گام ۳ + ۴ — جدول صنایع + امتیاز مرکب
# ──────────────────────────────────────────────────────────────────────────────
def build_industries(stocks: pd.DataFrame, total_mktcap: float) -> pd.DataFrame:
    g = stocks.groupby("صنعت")
    ind = g[IND_MEDIAN_COLS].median().round(2)
    ind["تعداد"] = g.size()
    ind["وزن_بازار٪"] = (g["ارزش_بازار_همت"].sum() / total_mktcap * 100).round(2)

    # دلاری٪ = درصد شرکت‌های صادرات‌محور (خالص ارزی مثبت) نسبت به شرکت‌های دارای داده
    def dollari(x):
        nf = x.dropna()
        return np.nan if len(nf) == 0 else round((nf > 0).mean() * 100)
    ind["دلاری٪"] = g["خالص_ارزی"].apply(dollari)

    ind = ind.reset_index()

    # امتیاز مرکب — فقط صنایع واجد شرایط (≥۴ شرکت)
    elig = ind[ind["تعداد"] >= MIN_STOCKS_FOR_SCORE].copy()
    z_value = -zscore(elig["PE"])                          # P/E پایین‌تر = بهتر (معکوس)
    z_growth = zscore(elig["رشد_سود"]).clip(-2, 2)         # مهار outlier در بازهٔ ±۲
    z_quality = zscore(elig["ROE"])
    z_margin = zscore(elig["حاشیه_خالص"])
    elig["امتیاز"] = ((z_value + z_growth + z_quality + z_margin) / 4).round(2)

    ind = ind.merge(elig[["صنعت", "امتیاز"]], on="صنعت", how="left")
    ind["رتبه"] = ind["امتیاز"].rank(ascending=False, method="min")

    ind = ind.sort_values(["امتیاز"], ascending=False, na_position="last").reset_index(drop=True)
    return ind[INDUSTRY_COLS]


# ──────────────────────────────────────────────────────────────────────────────
# گام ۵ — جدول سکتورها
# ──────────────────────────────────────────────────────────────────────────────
def build_sectors(stocks: pd.DataFrame, total_mktcap: float) -> pd.DataFrame:
    g = stocks.groupby("سکتور")
    sec = pd.DataFrame({
        "وزن٪": (g["ارزش_بازار_همت"].sum() / total_mktcap * 100).round(2),
        "PE": g["PE"].median().round(2),
        "ROE": g["ROE"].median().round(1),
        "رشد_سود": g["رشد_سود"].median().round(1),
        "بازده_1y": g["بازده_1y"].median().round(1),
        "تعداد": g.size(),
    }).reset_index().rename(columns={"index": "سکتور"})
    # مدیریت NaN در میانه (به‌جای کرش، None)
    sec = sec.where(pd.notna(sec), None)
    return sec[SECTOR_COLS].sort_values("وزن٪", ascending=False).reset_index(drop=True)


# ──────────────────────────────────────────────────────────────────────────────
# گام ۶ — نگاشت سناریو
# ──────────────────────────────────────────────────────────────────────────────
def build_scenario(ind: pd.DataFrame) -> pd.DataFrame:
    scored = ind.dropna(subset=["امتیاز"])
    bull_fx = scored[scored["دلاری٪"] >= 50].sort_values("امتیاز", ascending=False)   # برندهٔ تشدید
    bull_rial = scored[scored["دلاری٪"] < 50].sort_values("امتیاز", ascending=False)  # برندهٔ گشایش
    n = max(len(bull_fx), len(bull_rial))

    def pad(seq, k):
        seq = list(seq) + [None] * (k - len(seq))
        return seq

    return pd.DataFrame({
        "🔴 برنده تشدید (دلاری)": pad(bull_fx["صنعت"], n),
        "امتیاز_تشدید": pad(bull_fx["امتیاز"], n),
        "🟢 برنده گشایش (ریالی)": pad(bull_rial["صنعت"], n),
        "امتیاز_گشایش": pad(bull_rial["امتیاز"], n),
    })


# ──────────────────────────────────────────────────────────────────────────────
# شیت راهنما
# ──────────────────────────────────────────────────────────────────────────────
def build_guide(date_fa: str, stocks: pd.DataFrame) -> pd.DataFrame:
    pe_med = round(stocks["PE"].median(), 1)
    pb_med = round(stocks["PB"].median(), 1)
    roe_med = round(stocks["ROE"].median())
    rows = [
        ("تاریخ داده", f"{date_fa} (منبع: Pouya Finance / Screener Plus)"),
        ("", ""),
        ("شیت‌های فایل:", ""),
        ("۱) صنایع", "همهٔ صنایع + امتیاز مرکب + رتبه"),
        ("۲) سهام", "همهٔ سهام + شاخص‌ها"),
        ("۳) سکتورها", "سکتورهای اصلی"),
        ("۴) نگاشت سناریو", "دلاری (تشدید) / ریالی (گشایش)"),
        ("", ""),
        ("روش امتیاز مرکب:", "میانگین z-score چهار بُعد (وزن مساوی):"),
        ("  ارزش", "P/E پایین (معکوس)"),
        ("  رشد", "رشد سود خالص یک‌ساله (محدود به ±۲)"),
        ("  کیفیت", "ROE"),
        ("  حاشیه", "حاشیه سود خالص"),
        ("", ""),
        ("⚠ هشدارهای تحلیلی:", ""),
        ("بانک", "ROE بالا اغلب از تسعیر ارز/تجدید ارزیابی، نه عملیات پایدار"),
        ("سیمان", "P/B بالا = گران از نظر دارایی؛ فقط ROE توجیه‌گر"),
        ("نفتی/کود", "بازده گذشته عالی ولی حاشیه پایین؛ گذشته ≠ آینده"),
        ("کشاورزی/شیرینی", "رشد بالا ولی وزن کوچک = محدودیت نقدشوندگی"),
        ("خالص ارزی", "فقط شرکت‌های دارای داده؛ با احتیاط تفسیر شود"),
        ("صنایع <۴ شرکت", "امتیاز محاسبه نشده (آماری بی‌معنی)"),
        ("", ""),
        ("میانه بازار", f"P/E {pe_med} | P/B {pb_med} | ROE {roe_med}%"),
    ]
    return pd.DataFrame(rows, columns=[f"تحلیل کامل بورس ایران — {date_fa}", ""])


# ──────────────────────────────────────────────────────────────────────────────
# نوشتن + قالب‌بندیِ اکسل
# ──────────────────────────────────────────────────────────────────────────────
def write_excel(path, guide, industries, stocks, sectors, scenario):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    sheets = [
        ("راهنما", guide), ("صنایع", industries), ("سهام", stocks),
        ("سکتورها", sectors), ("نگاشت سناریو", scenario),
    ]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets:
            df.to_excel(writer, sheet_name=name, index=False)

        wb = writer.book
        navy = PatternFill("solid", fgColor="1E3A8A")
        green = PatternFill("solid", fgColor="C6EFCE")
        head_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        body_font = Font(name="Arial", size=10)

        for name, df in sheets:
            ws = writer.sheets[name]
            ws.sheet_view.rightToLeft = True          # نمایش راست‌به‌چپ
            ws.freeze_panes = "A2"                      # فریزِ ردیف هدر

            # فونت بدنه + هدرِ navy/سفید/توپر
            for row in ws.iter_rows():
                for c in row:
                    c.font = body_font
            for c in ws[1]:
                c.font = head_font
                c.fill = navy
                c.alignment = Alignment(horizontal="center", vertical="center")

            # هایلایت سبزِ ۵ صنعت برتر (فقط شیت صنایع)
            if name == "صنایع" and "رتبه" in df.columns:
                rank_idx = list(df.columns).index("رتبه")
                for r in range(len(df)):
                    rv = df.iloc[r, rank_idx]
                    if pd.notna(rv) and rv <= 5:
                        for c in ws[r + 2]:           # +2: هدر + ۰-اندیس
                            c.fill = green

            # عرضِ ستونِ خودکار
            for j, col in enumerate(df.columns, start=1):
                lengths = [len(str(col))] + [len(str(v)) for v in df.iloc[:, j - 1].tolist()]
                ws.column_dimensions[get_column_letter(j)].width = min(max(lengths) + 3, 48)


# ──────────────────────────────────────────────────────────────────────────────
# اجرا
# ──────────────────────────────────────────────────────────────────────────────
def main():
    raw_path = sys.argv[1] if len(sys.argv) > 1 else RAW_DEFAULT
    if not os.path.exists(raw_path):
        sys.exit(f"❌ فایل خام یافت نشد: {raw_path}")

    date_fa, date_tok = parse_date_token(raw_path)
    print(f"▶ فایل خام: {os.path.basename(raw_path)}  | تاریخ داده: {date_fa}")

    # خط‌لوله
    raw = pd.read_excel(raw_path, header=HEADER_ROW)
    total_mktcap = safe_num(raw[raw["Symbol"].notna()]["MarketCap"]).sum()

    stocks = build_stocks(raw_path)
    industries = build_industries(stocks, total_mktcap)
    sectors = build_sectors(stocks, total_mktcap)
    scenario = build_scenario(industries)
    guide = build_guide(date_fa, stocks)

    # نوشتنِ نسخهٔ تاریخ‌دار + کپیِ ثابتِ «جاری»
    dated = os.path.join(HERE, f"تحلیل_کامل_بورس_{date_tok}.xlsx")
    current = os.path.join(HERE, CURRENT_NAME)
    write_excel(dated, guide, industries, stocks, sectors, scenario)
    shutil.copyfile(dated, current)

    # گزارش خلاصه (برای راستی‌آزمایی)
    scored = industries.dropna(subset=["امتیاز"]).sort_values("امتیاز", ascending=False)
    top = scored.iloc[0]
    print(f"✓ صنایع: {len(industries)} ردیف | سهام: {len(stocks)} ردیف | سکتورها: {len(sectors)}")
    print(f"✓ رتبهٔ ۱: {top['صنعت']} (امتیاز {top['امتیاز']}) | میانه بازار P/E: {round(stocks['PE'].median(),2)}")
    print(f"✓ ساخته شد:\n   {dated}\n   {current}")


if __name__ == "__main__":
    main()
